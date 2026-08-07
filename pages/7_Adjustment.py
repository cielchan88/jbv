"""
⚙️ Adjustment - Manual Forecast Adjustment

Page untuk mengelola adjustment manual pada forecast.
- Pilih versi forecast yang ingin di-adjust
- Adjustment hanya berlaku untuk versi yang dipilih
- Lihat metadata lengkap (model, parameter, fitur)
- Download hasil adjustment

Author: APUVA Team
Date: 2025-11-11
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from utils.forecast_version import (
    list_forecast_versions,
    load_forecast_version,
    get_version_summary,
    format_timestamp
)
from utils.aggregation import recalculate_parents

# Page config
st.set_page_config(
    page_title="Adjustment",
    page_icon="⚙️",
    layout="wide"
)

# Header
st.title("⚙️ Adjustment")
st.markdown("Kelola adjustment manual pada forecast per versi.")

st.divider()

# Step 1: Select Forecast Version
st.markdown("### 📦 Pilih Versi Forecast")

# Get list of available versions
versions = list_forecast_versions()

if len(versions) == 0:
    st.warning("⚠️ Belum ada forecast yang tersimpan. Silahkan generate forecast terlebih dahulu di halaman **Lembar Kerja**")
    st.stop()

# Create selection options
version_options = {}
for v in versions:
    timestamp_formatted = format_timestamp(v['timestamp'])
    label = f"{timestamp_formatted} - {v['forecast_days']} hari - {v['model_mode']}"
    if v['config_name'] != '-':
        label += f" ({v['config_name']})"
    version_options[label] = v['version_id']

selected_version_label = st.selectbox(
    "Pilih versi forecast:",
    options=list(version_options.keys()),
    help="Pilih versi forecast yang ingin di-adjust. Setiap versi memiliki metadata lengkap."
)

selected_version_id = version_options[selected_version_label]

# Load selected version
with st.spinner(f"Loading forecast version {selected_version_id}..."):
    df_forecast, metadata = load_forecast_version(selected_version_id)

future_date_cols = metadata['future_date_cols']
leaf_nodes = metadata['leaf_nodes']
time_cols = metadata['time_cols']

st.success(f"✅ Loaded forecast version: **{selected_version_id}**")

# Show metadata summary
with st.expander("📊 Lihat Metadata Lengkap", expanded=False):
    summary = get_version_summary(selected_version_id)

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Forecast Days", summary.get('forecast_days', 0))
    with col2:
        st.metric("Total Leaf Nodes", summary.get('total_leaf_nodes', 0))
    with col3:
        st.metric("Model Mode", summary.get('model_mode', 'Unknown'))
    with col4:
        st.metric("File Size", f"{summary.get('file_size_kb', 0):.1f} KB")

    st.markdown("---")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**📅 Date Range**")
        st.write(f"- Last Historical: `{summary.get('last_historical_date', '-')}`")
        st.write(f"- First Forecast: `{summary.get('first_forecast_date', '-')}`")
        st.write(f"- Last Forecast: `{summary.get('last_forecast_date', '-')}`")

        st.markdown("**🔧 Configuration**")
        st.write(f"- Config Name: `{summary.get('config_name', '-')}`")
        st.write(f"- Cross-Series: `{'Yes' if summary.get('cross_series_used', False) else 'No'}`")
        if summary.get('cross_series_used', False):
            st.write(f"- Cross-Series Top K: `{summary.get('cross_series_top_k', 0)}`")
        st.write(f"- Holidays Count: `{summary.get('holidays_count', 0)}`")

    with col2:
        st.markdown("**🤖 Model Usage**")
        model_counts = summary.get('model_counts', {})
        if len(model_counts) > 0:
            for model_name, count in sorted(model_counts.items(), key=lambda x: x[1], reverse=True):
                st.write(f"- **{model_name}**: {count} nodes")
        else:
            st.write("No model info available")

st.divider()

# Initialize adjustments for this version
# Each version has its own adjustments stored separately
adjustment_key = f'adjustments_{selected_version_id}'
if adjustment_key not in st.session_state:
    st.session_state[adjustment_key] = []

# Step 2: Manage Adjustments
st.markdown("### 📝 Kelola Adjustment")

st.markdown("""
Gunakan tabel di bawah untuk menambah/edit/hapus adjustment.
- **Tambah baris baru**: Klik ➕ di pojok kanan tabel
- **Edit**: Klik cell yang ingin diubah
- **Hapus**: Centang baris, lalu klik 🗑️
""")

# Get leaf node info for better display
leaf_options = {}
leaf_options_reverse = {}
for leaf_id in leaf_nodes:
    leaf_row = df_forecast[df_forecast['Row_ID'] == leaf_id]
    if len(leaf_row) > 0:
        label = leaf_row['Row_Label'].values[0]
        display = f"{leaf_id} - {label}"
        leaf_options[leaf_id] = display
        leaf_options_reverse[display] = leaf_id

# Create editable dataframe from adjustments
if len(st.session_state[adjustment_key]) > 0:
    # Convert to DataFrame
    df_adjustments = pd.DataFrame(st.session_state[adjustment_key])
    # Add display label
    df_adjustments['component_display'] = df_adjustments['component_label']

    # Reorder columns for display
    display_df = df_adjustments[[
        'component_display', 'type', 'value',
        'start_date', 'end_date', 'note'
    ]].copy()
    display_df.columns = ['Komponen', 'Tipe', 'Nilai', 'Tanggal Mulai', 'Tanggal Akhir', 'Catatan']
else:
    # Empty DataFrame with correct structure
    display_df = pd.DataFrame(columns=[
        'Komponen', 'Tipe', 'Nilai', 'Tanggal Mulai', 'Tanggal Akhir', 'Catatan'
    ])

# Editable table
edited_df = st.data_editor(
    display_df,
    num_rows="dynamic",  # Allow add/delete rows
    use_container_width=True,
    column_config={
        "Komponen": st.column_config.SelectboxColumn(
            "Komponen",
            options=list(leaf_options.values()),
            required=True,
            width="large"
        ),
        "Tipe": st.column_config.SelectboxColumn(
            "Tipe",
            options=["Tambah (+)", "Kurang (-)"],
            required=True,
            width="small"
        ),
        "Nilai": st.column_config.NumberColumn(
            "Nilai",
            min_value=0.0,
            format="%.0f",
            required=True,
            width="small"
        ),
        "Tanggal Mulai": st.column_config.SelectboxColumn(
            "Tanggal Mulai",
            options=future_date_cols,
            required=True,
            width="medium"
        ),
        "Tanggal Akhir": st.column_config.SelectboxColumn(
            "Tanggal Akhir",
            options=future_date_cols,
            required=True,
            width="medium"
        ),
        "Catatan": st.column_config.TextColumn(
            "Catatan",
            width="large"
        )
    },
    hide_index=True,
    key="adjustments_editor"
)

# Sync edited_df back to session state
# Convert back to adjustment records
new_adjustments = []
for idx, row in edited_df.iterrows():
    # Skip incomplete rows
    if pd.isna(row['Komponen']) or pd.isna(row['Nilai']) or row['Nilai'] <= 0:
        continue

    # Get component ID from display name
    component_display = row['Komponen']
    component_id = leaf_options_reverse.get(component_display, None)

    if component_id is None:
        continue

    adjustment_record = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'component_id': component_id,
        'component_label': component_display,
        'type': row['Tipe'] if not pd.isna(row['Tipe']) else "Tambah (+)",
        'value': float(row['Nilai']),
        'start_date': row['Tanggal Mulai'] if not pd.isna(row['Tanggal Mulai']) else future_date_cols[0],
        'end_date': row['Tanggal Akhir'] if not pd.isna(row['Tanggal Akhir']) else future_date_cols[-1],
        'note': row['Catatan'] if not pd.isna(row['Catatan']) else '-',
        'applied': False
    }
    new_adjustments.append(adjustment_record)

# Update session state
st.session_state[adjustment_key] = new_adjustments

# Show real-time preview if adjustments exist
if len(st.session_state[adjustment_key]) > 0:
    st.divider()
    st.markdown("### 👁️ Preview Hasil Adjustment")

    # Calculate preview with adjustments
    df_preview = df_forecast.copy()

    for adj in st.session_state[adjustment_key]:
        component_id = adj['component_id']
        start_date = adj['start_date']
        end_date = adj['end_date']
        value = adj['value']
        adj_type = adj['type']

        # Get date range
        date_range = [d for d in future_date_cols if start_date <= d <= end_date]

        # Apply adjustment to leaf node
        component_idx = df_preview[df_preview['Row_ID'] == component_id].index
        if len(component_idx) > 0:
            for date_col in date_range:
                current_val = df_preview.at[component_idx[0], date_col]
                if adj_type == "Tambah (+)":
                    new_val = current_val + value
                else:
                    new_val = current_val - value
                df_preview.at[component_idx[0], date_col] = round(new_val, 2)

    # RECALCULATE PARENTS after all leaf adjustments
    # This ensures A.1 = sum(A.1.a + A.1.b + ...) and A = sum(A.1 + A.2 + ...)
    df_preview = recalculate_parents(df_preview, future_date_cols)

    # Show full adjusted table (like in Lembar Kerja)
    st.markdown("**Tabel lengkap dengan nilai yang sudah di-adjust:**")

    # Show last 10 historical dates + all forecast dates
    preview_cols = ['Row_ID', 'Row_Label', 'Level'] + time_cols[-10:] + future_date_cols
    preview_df_display = df_preview[preview_cols]

    # Style forecast columns with red background (adjusted values)
    def highlight_forecast_cols(s):
        styles = pd.DataFrame('', index=s.index, columns=s.columns)
        # Red background for forecast columns
        for col in future_date_cols:
            if col in styles.columns:
                styles[col] = 'background-color: #ffcccc'
        return styles

    styled_preview = preview_df_display.style.apply(highlight_forecast_cols, axis=None)
    st.dataframe(styled_preview, use_container_width=True, height=600)

    # Get unique components from adjustments
    adjusted_components = list(set([adj['component_id'] for adj in st.session_state[adjustment_key]]))

    st.success(f"✅ Total {len(st.session_state[adjustment_key])} adjustment telah ditambahkan untuk versi **{selected_version_id}**.")
    st.info(f"📊 Komponen yang di-adjust: **{', '.join(adjusted_components)}**")

    # Show comparison details per component in expander
    with st.expander("📋 Lihat Detail Perubahan Per Komponen", expanded=False):
        for comp_id in adjusted_components:
            st.markdown(f"**{comp_id}**")

            # Get original and adjusted values
            orig_row = df_forecast[df_forecast['Row_ID'] == comp_id]
            adj_row = df_preview[df_preview['Row_ID'] == comp_id]

            if len(orig_row) > 0 and len(adj_row) > 0:
                # Get affected dates
                affected_dates = []
                for adj in st.session_state[adjustment_key]:
                    if adj['component_id'] == comp_id:
                        affected_dates.extend([d for d in future_date_cols if adj['start_date'] <= d <= adj['end_date']])
                affected_dates = sorted(list(set(affected_dates)))

                # Create comparison dataframe
                if len(affected_dates) > 0:
                    comparison_data = {
                        'Tanggal': affected_dates,
                        'Original': [orig_row.iloc[0][d] for d in affected_dates],
                        'Adjusted': [adj_row.iloc[0][d] for d in affected_dates],
                        'Diff': [adj_row.iloc[0][d] - orig_row.iloc[0][d] for d in affected_dates]
                    }
                    comparison_df = pd.DataFrame(comparison_data)
                    comparison_df['Diff'] = comparison_df['Diff'].apply(lambda x: f"{x:+,.0f}")

                    st.dataframe(comparison_df, use_container_width=True, hide_index=True)
                    st.markdown("---")

else:
    st.info("📝 Belum ada adjustment untuk versi ini. Klik ➕ di pojok kanan tabel untuk menambah adjustment baru.")

# Download Section
st.divider()
st.markdown("### 📥 Download Forecast")

col1, col2 = st.columns([2, 1])

with col1:
    # Download adjusted forecast
    if len(st.session_state[adjustment_key]) > 0:
        from io import BytesIO

        # Apply adjustments to get final dataframe
        df_adjusted = df_forecast.copy()

        for adj in st.session_state[adjustment_key]:
            component_id = adj['component_id']
            start_date = adj['start_date']
            end_date = adj['end_date']
            value = adj['value']
            adj_type = adj['type']

            date_range = [d for d in future_date_cols if start_date <= d <= end_date]
            component_idx = df_adjusted[df_adjusted['Row_ID'] == component_id].index

            if len(component_idx) > 0:
                for date_col in date_range:
                    current_val = df_adjusted.at[component_idx[0], date_col]
                    if adj_type == "Tambah (+)":
                        new_val = current_val + value
                    else:
                        new_val = current_val - value
                    df_adjusted.at[component_idx[0], date_col] = round(new_val, 2)

        # RECALCULATE PARENTS after all leaf adjustments
        # This ensures totals are consistent: A.1 = sum(children), A = sum(A.1 + A.2 + ...)
        df_adjusted = recalculate_parents(df_adjusted, future_date_cols)

        # Create Excel with metadata sheet
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Adjusted Forecast
            df_adjusted.to_excel(writer, index=False, sheet_name='SDV Forecast Adjusted')

            # Format forecast columns with red background
            from openpyxl.styles import PatternFill
            worksheet = writer.sheets['SDV Forecast Adjusted']

            # Red fill for forecast columns
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

            # Find column indices for forecast dates
            header_row = list(df_adjusted.columns)
            forecast_col_indices = []
            for col_name in future_date_cols:
                if col_name in header_row:
                    col_idx = header_row.index(col_name) + 1  # openpyxl uses 1-based indexing
                    forecast_col_indices.append(col_idx)

            # Apply red background to forecast columns (header + all data rows)
            for col_idx in forecast_col_indices:
                # Header row
                worksheet.cell(row=1, column=col_idx).fill = red_fill
                # Data rows
                for row_idx in range(2, len(df_adjusted) + 2):  # +2 because 1=header, data starts at 2
                    worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

            # Sheet 2: Adjustment Log
            if len(st.session_state[adjustment_key]) > 0:
                adjustment_log = []
                for adj in st.session_state[adjustment_key]:
                    adjustment_log.append({
                        'Component ID': adj['component_id'],
                        'Component Label': adj['component_label'],
                        'Type': adj['type'],
                        'Value': adj['value'],
                        'Start Date': adj['start_date'],
                        'End Date': adj['end_date'],
                        'Note': adj['note']
                    })
                df_adj_log = pd.DataFrame(adjustment_log)
                df_adj_log.to_excel(writer, index=False, sheet_name='Adjustment Log')

            # Sheet 3: Metadata
            summary = get_version_summary(selected_version_id)
            metadata_rows = []
            metadata_rows.append(['Version ID', selected_version_id])
            metadata_rows.append(['Timestamp', format_timestamp(summary.get('timestamp', ''))])
            metadata_rows.append(['Forecast Days', summary.get('forecast_days', 0)])
            metadata_rows.append(['Model Mode', summary.get('model_mode', '-')])
            metadata_rows.append(['Config Name', summary.get('config_name', '-')])
            metadata_rows.append(['Total Leaf Nodes', summary.get('total_leaf_nodes', 0)])
            metadata_rows.append(['Cross-Series Used', 'Yes' if summary.get('cross_series_used', False) else 'No'])
            metadata_rows.append(['Cross-Series Top K', summary.get('cross_series_top_k', 0)])
            metadata_rows.append(['Holidays Count', summary.get('holidays_count', 0)])
            metadata_rows.append(['Last Historical Date', summary.get('last_historical_date', '-')])
            metadata_rows.append(['First Forecast Date', summary.get('first_forecast_date', '-')])
            metadata_rows.append(['Last Forecast Date', summary.get('last_forecast_date', '-')])
            metadata_rows.append(['', ''])
            metadata_rows.append(['Model Usage', ''])
            for model_name, count in summary.get('model_counts', {}).items():
                metadata_rows.append([model_name, count])

            df_metadata = pd.DataFrame(metadata_rows, columns=['Parameter', 'Value'])
            df_metadata.to_excel(writer, index=False, sheet_name='Metadata')

            # Sheet 4: Series Summary (SAME AS Lembar_Kerja.py)
            series_details = summary.get('series_details', [])
            if len(series_details) > 0:
                series_rows = []
                for detail in series_details:
                    series_rows.append({
                        'Row_ID': detail.get('row_id', '-'),
                        'Row_Label': detail.get('row_label', '-'),
                        'Model': detail.get('model', '-'),
                        'Uses_Features': 'Yes' if detail.get('uses_features', False) else 'No',
                        'Features_Selected': detail.get('total_features_selected', 0),
                        'Features_Generated': detail.get('total_features_generated', 0)
                    })

                df_series_summary = pd.DataFrame(series_rows)
                df_series_summary.to_excel(writer, index=False, sheet_name='Series Summary')

            # Sheet 5: Feature Details (SAME AS Lembar_Kerja.py)
            # ALL 25 features per ML model series with correlation scores
            feature_detail_rows = []
            for detail in series_details:
                if detail.get('uses_features', False):
                    row_id = detail.get('row_id', '-')
                    row_label = detail.get('row_label', '-')
                    feature_scores = detail.get('feature_scores_all', {})

                    if feature_scores:
                        for rank, (feature_name, score) in enumerate(feature_scores.items(), start=1):
                            feature_detail_rows.append({
                                'Row_ID': row_id,
                                'Row_Label': row_label,
                                'Rank': rank,
                                'Feature_Name': feature_name,
                                'Correlation_Score': round(score, 4)
                            })

            if len(feature_detail_rows) > 0:
                df_feature_details = pd.DataFrame(feature_detail_rows)
                df_feature_details.to_excel(writer, index=False, sheet_name='Feature Details')

            # Sheet 6: Feature Descriptions Reference
            feature_descriptions = [
                # Historis
                {'Tipe': 'Historis', 'Pattern': 'lag_N', 'Penjelasan': 'Nilai aktual N hari kerja yang lalu, menangkap pola jangka pendek'},
                # Statistik
                {'Tipe': 'Statistik', 'Pattern': 'rolling_mean_N', 'Penjelasan': 'Rata-rata bergerak N hari terakhir, menghaluskan fluktuasi harian'},
                {'Tipe': 'Statistik', 'Pattern': 'rolling_std_N', 'Penjelasan': 'Standar deviasi N hari terakhir, mengukur tingkat volatilitas'},
                {'Tipe': 'Statistik', 'Pattern': 'rolling_max_N', 'Penjelasan': 'Nilai tertinggi dalam N hari terakhir, mendeteksi puncak lokal'},
                {'Tipe': 'Statistik', 'Pattern': 'rolling_min_N', 'Penjelasan': 'Nilai terendah dalam N hari terakhir, mendeteksi lembah lokal'},
                {'Tipe': 'Statistik', 'Pattern': 'ewm_N', 'Penjelasan': 'Rata-rata eksponensial N hari, memberi bobot lebih pada data terbaru'},
                {'Tipe': 'Statistik', 'Pattern': 'ewm_std_N', 'Penjelasan': 'Volatilitas eksponensial N hari, lebih responsif terhadap perubahan terkini'},
                # Teknikal
                {'Tipe': 'Teknikal', 'Pattern': 'bb_middle_N', 'Penjelasan': 'Garis tengah Bollinger Band (SMA N hari), basis untuk mengukur deviasi'},
                {'Tipe': 'Teknikal', 'Pattern': 'bb_upper_N', 'Penjelasan': 'Batas atas Bollinger Band, sinyal potensi overbought'},
                {'Tipe': 'Teknikal', 'Pattern': 'bb_lower_N', 'Penjelasan': 'Batas bawah Bollinger Band, sinyal potensi oversold'},
                {'Tipe': 'Teknikal', 'Pattern': 'bb_width_N', 'Penjelasan': 'Lebar Bollinger Band, indikator ekspansi/kontraksi volatilitas'},
                {'Tipe': 'Teknikal', 'Pattern': 'macd', 'Penjelasan': 'Moving Average Convergence Divergence, mengukur momentum dan arah tren'},
                {'Tipe': 'Teknikal', 'Pattern': 'macd_signal', 'Penjelasan': 'Signal line MACD, digunakan untuk sinyal buy/sell crossover'},
                {'Tipe': 'Teknikal', 'Pattern': 'rsi_N', 'Penjelasan': 'Relative Strength Index N hari, mengukur kekuatan pergerakan (0-100)'},
                {'Tipe': 'Teknikal', 'Pattern': 'momentum_N', 'Penjelasan': 'Selisih nilai dengan N hari lalu, mengukur kecepatan perubahan'},
                {'Tipe': 'Teknikal', 'Pattern': 'rate_of_change_N', 'Penjelasan': 'Persentase perubahan dari N hari lalu, momentum relatif'},
                # Kalender
                {'Tipe': 'Kalender', 'Pattern': 'day_of_week', 'Penjelasan': 'Hari dalam minggu (0=Senin s.d. 4=Jumat), menangkap pola mingguan'},
                {'Tipe': 'Kalender', 'Pattern': 'day_of_month', 'Penjelasan': 'Tanggal dalam bulan (1-31), menangkap pola awal/akhir bulan'},
                {'Tipe': 'Kalender', 'Pattern': 'month', 'Penjelasan': 'Bulan dalam tahun (1-12), menangkap pola musiman'},
                {'Tipe': 'Kalender', 'Pattern': 'quarter', 'Penjelasan': 'Kuartal dalam tahun (1-4), menangkap pola kuartalan'},
                {'Tipe': 'Kalender', 'Pattern': 'week_of_year', 'Penjelasan': 'Minggu ke-N dalam tahun (1-52), menangkap siklus tahunan'},
                {'Tipe': 'Kalender', 'Pattern': 'is_month_start', 'Penjelasan': 'Indikator awal bulan (1/0), menangkap efek turn-of-month'},
                {'Tipe': 'Kalender', 'Pattern': 'is_month_end', 'Penjelasan': 'Indikator akhir bulan (1/0), menangkap efek window dressing'},
                {'Tipe': 'Kalender', 'Pattern': 'is_quarter_start', 'Penjelasan': 'Indikator awal kuartal (1/0), menangkap efek rebalancing'},
                {'Tipe': 'Kalender', 'Pattern': 'is_quarter_end', 'Penjelasan': 'Indikator akhir kuartal (1/0), menangkap efek reporting'},
                {'Tipe': 'Kalender', 'Pattern': 'is_holiday', 'Penjelasan': 'Indikator hari libur (1/0), menangkap dampak libur terhadap transaksi'},
                {'Tipe': 'Kalender', 'Pattern': 'days_to_holiday', 'Penjelasan': 'Jumlah hari menuju libur terdekat, antisipasi pra-libur'},
                {'Tipe': 'Kalender', 'Pattern': 'days_from_holiday', 'Penjelasan': 'Jumlah hari sejak libur terakhir, efek pasca-libur'},
                # Eksternal
                {'Tipe': 'Eksternal', 'Pattern': 'news_count', 'Penjelasan': 'Jumlah berita ekonomi/keuangan harian dari Trading Economics'},
                {'Tipe': 'Eksternal', 'Pattern': 'sentiment_finbert', 'Penjelasan': 'Skor sentimen berita dari model FinBERT (0-1)'},
                {'Tipe': 'Eksternal', 'Pattern': 'sentiment_bertmulti', 'Penjelasan': 'Skor sentimen dari BERT Multilingual untuk berita Indonesia'},
                {'Tipe': 'Eksternal', 'Pattern': 'sentiment_finbert_weighted', 'Penjelasan': 'Sentimen FinBERT tertimbang confidence score'},
                {'Tipe': 'Eksternal', 'Pattern': 'oil_price', 'Penjelasan': 'Harga minyak mentah global (USD/barrel)'},
                {'Tipe': 'Eksternal', 'Pattern': 'usd_idr', 'Penjelasan': 'Kurs USD/IDR dari pasar valuta asing'},
                {'Tipe': 'Eksternal', 'Pattern': 'gold', 'Penjelasan': 'Harga emas global (USD/oz)'},
                {'Tipe': 'Eksternal', 'Pattern': 'us_treasury', 'Penjelasan': 'Yield US Treasury 10Y, indikator suku bunga global'},
                # Series Lain
                {'Tipe': 'Series Lain', 'Pattern': 'A.x.x.x / B.x.x / C.x.x', 'Penjelasan': 'Data dari series lain yang berkorelasi tinggi dengan target (cross-series features)'},
                {'Tipe': 'Series Lain', 'Pattern': 'ext_[Series_ID]', 'Penjelasan': 'Lag/rolling dari series lain yang berkorelasi tinggi'},
            ]
            df_feature_desc = pd.DataFrame(feature_descriptions)
            df_feature_desc.to_excel(writer, index=False, sheet_name='Feature Descriptions')

        buffer.seek(0)

        st.download_button(
            label="📥 Download Forecast Adjusted (dengan Metadata Lengkap)",
            data=buffer,
            file_name=f"sdv_forecast_adjusted_{selected_version_id}_{datetime.now().strftime('%H%M')}.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type="primary",
            use_container_width=True
        )

        st.caption(f"""
        💡 File berisi 6 sheet:
        - **Adjusted Forecast**: Data forecast yang sudah di-adjust manual
        - **Adjustment Log**: Detail semua adjustment yang diterapkan
        - **Metadata**: Informasi versi, model usage, dan parameters
        - **Series Summary**: Ringkasan per-series (model apa, berapa fitur yang dipakai)
        - **Feature Details**: SEMUA 25 fitur terpilih per series ML dengan skor korelasi Spearman
        - **Feature Descriptions**: Penjelasan setiap tipe fitur (Historis, Statistik, Teknikal, dll)
        """)
    else:
        st.info("📝 Belum ada adjustment. Tambahkan adjustment terlebih dahulu untuk download versi adjusted.")

with col2:
    # Button to clear all adjustments
    if st.button("🗑️ Hapus Semua Adjustment", type="secondary", disabled=len(st.session_state[adjustment_key])==0, use_container_width=True):
        st.session_state[adjustment_key] = []
        st.rerun()

# Footer
st.divider()
st.caption("🔮 Adjustment tersimpan per versi. Generate forecast baru akan membuat versi baru (adjustment tidak carry over).")
