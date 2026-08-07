import streamlit as st
import pandas as pd
import numpy as np
import warnings
import os
import json
from datetime import datetime
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Lembar Kerja - JBV Dashboard", layout="wide")

# Import utils
from utils import load_holidays, generate_business_dates
from utils.data_loader import load_etl_output, parse_children
from utils.forecasting import forecast_single_series
from utils.forecast_version import save_forecast_version, list_forecast_versions

st.title("📋 Lembar Kerja")
st.markdown("Generate forecast untuk semua komponen dan download hasilnya.")

st.divider()

st.subheader("🔮 Forecast Hari Kedepan")

# Input for number of forecast days
col1, col2 = st.columns([1, 1])

with col1:
    forecast_days = st.number_input(
        "Jumlah hari forecast kedepan:",
        min_value=1,
        max_value=365,
        value=30,
        step=1,
        help="Tentukan berapa hari kedepan yang ingin diprediksi (default: 30 hari)"
    )

with col2:
    model_mode = st.selectbox(
        "Mode pemilihan model:",
        options=["Custom (Hasil Evaluasi)", "Stacking", "APUVA", "Prophet", "RandomForest", "LightGBM", "XGBoost", "AutoARIMA", "VAR"],
        help="Custom: setiap leaf pakai model terbaiknya dari evaluasi. Single Model: semua leaf pakai 1 model yang sama"
    )

if model_mode == "Custom (Hasil Evaluasi)":
    st.markdown(f"""
    Generate prediksi **{forecast_days} hari** kedepan untuk **semua leaf nodes**.
    Model yang digunakan berdasarkan **konfigurasi terbaik** dari halaman **Evaluation**.
    """)
else:
    st.markdown(f"""
    Generate prediksi **{forecast_days} hari** kedepan untuk **semua leaf nodes**.
    Semua leaf nodes akan menggunakan model **{model_mode}**.
    """)

# Load ETL output data (NOT raw processed data!)
@st.cache_data(ttl=300)  # Cache for 5 minutes, then reload to get fresh data
def load_forecast_data():
    df, metadata_cols, time_cols = load_etl_output()

    # Get leaf nodes (nodes without children - parse JSON strings)
    leaf_nodes = df[df['Children'].apply(lambda x: len(parse_children(x))) == 0]['Row_ID'].tolist()

    return df, metadata_cols, time_cols, leaf_nodes

df, metadata_cols, time_cols, leaf_nodes = load_forecast_data()

# Load or set model configuration based on mode
if model_mode == "Custom (Hasil Evaluasi)":
    config_dir = 'model_configs'

    # Get list of available configurations
    available_configs = []
    if os.path.exists(config_dir):
        for filename in os.listdir(config_dir):
            if filename.endswith('.json'):
                try:
                    filepath = os.path.join(config_dir, filename)
                    with open(filepath, 'r') as f:
                        config_data = json.load(f)
                        if 'metadata' in config_data:
                            available_configs.append({
                                'filepath': filepath,
                                'filename': filename,
                                'name': config_data['metadata']['name'],
                                'date': config_data['metadata']['date'],
                                'test_size': config_data['metadata'].get('test_size', 'N/A'),
                                'metric': config_data['metadata'].get('metric', 'WAPE'),
                                'data': config_data
                            })
                except:
                    pass

    if len(available_configs) > 0:
        # Sort by date (newest first)
        available_configs = sorted(available_configs, key=lambda x: x['date'], reverse=True)

        # Select configuration
        config_options = {f"{cfg['name']} ({cfg['date']}) - Test: {cfg['test_size']}%": cfg for cfg in available_configs}

        selected_config_name = st.selectbox(
            "Pilih Konfigurasi Model:",
            options=list(config_options.keys()),
            help="Pilih konfigurasi model yang telah disimpan dari halaman Evaluasi"
        )

        selected_config = config_options[selected_config_name]
        model_config = selected_config['data']['models']

        st.success(f"✅ Konfigurasi '{selected_config['name']}' berhasil dimuat untuk {len(model_config)} leaf nodes")

        # Show summary of model usage
        model_counts = {}
        for model in model_config.values():
            model_counts[model] = model_counts.get(model, 0) + 1

        col1, col2, col3, col4, col5, col6, col7, col8 = st.columns(8)
        cols = [col1, col2, col3, col4, col5, col6, col7, col8]
        model_order = ['Stacking', 'APUVA', 'Prophet', 'RandomForest', 'LightGBM', 'XGBoost', 'AutoARIMA', 'VAR']

        for idx, model_name in enumerate(model_order):
            if model_name in model_counts:
                with cols[idx]:
                    st.metric(model_name, f"{model_counts[model_name]} nodes")
    else:
        st.warning("⚠️ Konfigurasi model belum tersedia. Silakan jalankan **Evaluasi** terlebih dahulu.")
        st.info("💡 Untuk sementara, semua leaf nodes akan menggunakan **Prophet** sebagai model default.")
        model_config = {leaf_id: 'Prophet' for leaf_id in leaf_nodes}
else:
    # Single model mode - all leafs use the same model
    single_model = model_mode  # e.g., "ElasticNet", "ARIMA", etc.
    model_config = {leaf_id: single_model for leaf_id in leaf_nodes}
    st.info(f"📌 Mode Single Model: Semua {len(leaf_nodes)} leaf nodes akan menggunakan **{single_model}**")

st.info(f"📊 Total **{len(leaf_nodes)}** leaf nodes akan diprediksi untuk **{forecast_days} hari** kedepan")

# Load holidays for business date generation
holidays = load_holidays()

# Pre-calculate cross-series correlations for ALL leaf nodes
# This ensures consistency with predictive-new.py
@st.cache_data(ttl=300, show_spinner=False)  # Cache for 5 minutes, silent execution
def prepare_cross_series_data(df, leaf_nodes, time_cols):
    """
    Prepare cross-series correlation data for all leaf nodes

    Uses 2019+ data for ML models with external features integration

    NOTE: This function is SILENT - no Streamlit messages inside to avoid duplication
    """
    from utils.feature_engineering_optimized import (
        calculate_series_correlations,
        select_top_correlated_series,
        prepare_external_series_data
    )
    import pandas as pd

    # Filter time_cols to 2019+ for external features alignment
    time_cols_ml = [col for col in time_cols if pd.to_datetime(col) >= pd.Timestamp('2019-01-01')]

    cross_series_map = {}

    for leaf_id in leaf_nodes:
        # 1. Get all other leaf nodes (exclude current series)
        candidate_series = [lid for lid in leaf_nodes if lid != leaf_id]

        # 2. Calculate correlations with ALL other leaf nodes (using 2019+ data)
        correlations = calculate_series_correlations(df, leaf_id, candidate_series, time_cols_ml)

        # 3. Select top 30 correlated series
        top_30_series = select_top_correlated_series(correlations, top_k=30)

        # 4. Prepare external series data (cross-series only) - using 2019+ data
        cross_series_only = prepare_external_series_data(df, top_30_series, time_cols_ml)

        # 5. Merge with external features from Excel (Sentiment, Oil, USD_IDR, etc)
        from utils.external_loader import load_and_merge_external_features
        external_series_data = load_and_merge_external_features(cross_series_only)

        cross_series_map[leaf_id] = external_series_data

    return cross_series_map

# Check which models need cross-series features (only ML models + VAR)
ml_models = {'RandomForest', 'XGBoost', 'LightGBM', 'Stacking', 'VAR'}
needs_cross_series = any(model_config.get(leaf_id, '').upper() in [m.upper() for m in ml_models] for leaf_id in leaf_nodes)

# Only calculate cross-series data if ML models are used
if needs_cross_series:
    # Get list of leaf nodes that use ML models
    ml_leaf_nodes = [leaf_id for leaf_id in leaf_nodes if model_config.get(leaf_id, '').upper() in [m.upper() for m in ml_models]]

    with st.spinner(f"🔍 Calculating cross-series correlations for {len(ml_leaf_nodes)} ML model leaf nodes..."):
        cross_series_map = prepare_cross_series_data(df, ml_leaf_nodes, time_cols)

        # Check if external features were loaded (check first leaf node)
        first_leaf = list(cross_series_map.values())[0] if cross_series_map else {}
        external_count = sum(1 for k in first_leaf.keys() if not k.startswith('A') and not k.startswith('B') and not k.startswith('C'))

        st.success(f"✅ Cross-series correlations calculated for {len(ml_leaf_nodes)} ML model leaf nodes")
        if external_count > 0:
            st.success(f"✅ External features integrated: {external_count} features from Trading Economics")
else:
    cross_series_map = {}
    st.info("ℹ️ No ML models used - skipping cross-series correlation calculation")

# Button to run forecast
if st.button(f"🚀 Generate Forecast {forecast_days} Hari untuk Semua Leaf Nodes", type="primary"):

    # Fungsi untuk melakukan forecast 1 leaf node dengan model yang dipilih
    # USING CENTRALIZED FORECASTING for 100% consistency with predictive-new.py
    def forecast_single_leaf(row_id, values, dates, model_type, n_days, expected_dates, external_series):
        try:
            # Use centralized forecasting function with cross-series features
            forecast_values, forecast_dates_list = forecast_single_series(
                dates=dates,
                values=values,
                model_name=model_type,
                n_days=n_days,
                holidays=holidays,
                external_series=external_series,
                row_id=row_id  # Pass row_id for APUVA sentiment factors
            )

            # DEBUG: Check if forecast dates match expected dates
            forecast_dates_str = [d.strftime('%Y-%m-%d') for d in forecast_dates_list]
            if forecast_dates_str != expected_dates:
                st.warning(f"⚠️ Date mismatch for {row_id}: expected {len(expected_dates)}, got {len(forecast_dates_str)}")

            return np.array(forecast_values), forecast_dates_list
        except Exception as e:
            st.warning(f"⚠️ Forecast error for {row_id} with {model_type}: {str(e)}")
            # Fallback to last value
            last_val = values[-1] if len(values) > 0 and not np.isnan(values[-1]) else 0
            return np.full(len(expected_dates), last_val), expected_dates

    # Fungsi untuk agregasi hierarki
    def aggregate_hierarchy(df_wide, forecast_dates):
        """
        Aggregate calculated nodes based on hierarchy structure
        Following logic from 003.r
        """
        df_result = df_wide.copy()

        # Define aggregation rules based on 003.r logic (ORDER MATTERS!)
        # Must aggregate from bottom to top of hierarchy
        aggregation_order = [
            # Level 3 - deepest children
            ('A.1.c', ['A.1.c.1', 'A.1.c.2', 'A.1.c.3', 'A.1.c.4', 'A.1.c.5', 'A.1.c.6', 'A.1.c.7', 'A.1.c.8']),
            ('B.d', ['B.d.1', 'B.d.2', 'B.d.3', 'B.d.4', 'B.d.5', 'B.d.6', 'B.d.7']),
            ('C.e', ['C.e.1', 'C.e.2', 'C.e.3', 'C.e.4', 'C.e.5', 'C.e.6']),
            # Level 2
            ('A.1', ['A.1.a', 'A.1.b', 'A.1.c']),
            ('A.0', ['A.0.a', 'A.0.b', 'A.0.c', 'A.0.d', 'A.0.e', 'A.0.f', 'A.0.g', 'A.0.h', 'A.0.i', 'A.0.j']),
            ('A.2.f', ['A.2.f.1', 'A.2.f.2', 'A.2.f.3', 'A.2.f.4', 'A.2.f.5']),
            ('A.2', ['A.2.a', 'A.2.b', 'A.2.c', 'A.2.d', 'A.2.e', 'A.2.f']),
            ('B', ['B.a', 'B.b', 'B.c', 'B.d']),
            ('C', ['C.a', 'C.b', 'C.c', 'C.d', 'C.e']),
            # Level 1
            ('A', ['A.1', 'A.2']),  # Line 151 in 003.r: A = A.1 + A.2
        ]

        # Helper function to get value safely
        def get_val(row_id, date_col):
            rows = df_result[df_result['Row_ID'] == row_id]
            if len(rows) > 0:
                return rows[date_col].values[0]
            return 0

        # Helper function to set value using .at
        def set_val(row_id, date_col, value):
            idx = df_result[df_result['Row_ID'] == row_id].index
            if len(idx) > 0:
                df_result.at[idx[0], date_col] = value

        # Special calculations for A.2 sub-components (round to 2 decimals)
        # ONLY if using FULL version (with A.0, A.1.c.x, etc.)
        has_a0 = len(df_result[df_result['Row_ID'] == 'A.0']) > 0

        if has_a0:
            # FULL VERSION: Calculate A.2.x from A.0.x - A.1.x
            for date_col in forecast_dates:
                # A.2.a = A.0.a
                set_val('A.2.a', date_col, round(get_val('A.0.a', date_col), 2))

                # A.2.d = A.0.b - A.1.a
                set_val('A.2.d', date_col, round(get_val('A.0.b', date_col) - get_val('A.1.a', date_col), 2))

                # A.2.b = A.0.i - A.1.c.7
                set_val('A.2.b', date_col, round(get_val('A.0.i', date_col) - get_val('A.1.c.7', date_col), 2))

                # A.2.c = A.0.e - A.1.c.4
                set_val('A.2.c', date_col, round(get_val('A.0.e', date_col) - get_val('A.1.c.4', date_col), 2))

                # A.2.e = A.0.f - A.1.b
                set_val('A.2.e', date_col, round(get_val('A.0.f', date_col) - get_val('A.1.b', date_col), 2))

                # A.2.f components
                set_val('A.2.f.1', date_col, round(get_val('A.0.c', date_col) - get_val('A.1.c.2', date_col), 2))
                set_val('A.2.f.2', date_col, round(get_val('A.0.d', date_col) - get_val('A.1.c.3', date_col), 2))
                set_val('A.2.f.3', date_col, round(get_val('A.0.g', date_col) - get_val('A.1.c.5', date_col), 2))
                set_val('A.2.f.4', date_col, round(get_val('A.0.h', date_col) - get_val('A.1.c.6', date_col), 2))
                set_val('A.2.f.5', date_col, round(get_val('A.0.j', date_col) - get_val('A.1.c.8', date_col), 2))
        # else: SIMPLE VERSION - A.2.x already forecasted as leaf nodes, no calculation needed

        # Aggregate using sum logic IN ORDER
        for parent, children in aggregation_order:
            parent_idx = df_result[df_result['Row_ID'] == parent].index
            if len(parent_idx) > 0:
                parent_idx = parent_idx[0]

                # Check if all children exist (for SIMPLE vs FULL compatibility)
                existing_children = [c for c in children if len(df_result[df_result['Row_ID'] == c]) > 0]

                if len(existing_children) > 0:
                    for date_col in forecast_dates:
                        total = 0
                        for child_id in existing_children:
                            child_val = get_val(child_id, date_col)
                            total += child_val

                        df_result.at[parent_idx, date_col] = round(total, 2)

        # D = A + B + C (top level)
        d_idx = df_result[df_result['Row_ID'] == 'D'].index
        if len(d_idx) > 0:
            d_idx = d_idx[0]
            for date_col in forecast_dates:
                a_val = get_val('A', date_col)
                b_val = get_val('B', date_col)
                c_val = get_val('C', date_col)
                d_val = a_val + b_val + c_val
                df_result.at[d_idx, date_col] = round(d_val, 2)

        return df_result

    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()

    # Copy original dataframe
    df_forecast = df.copy()
    df_forecast = df_forecast.drop('Children', axis=1)

    # Generate future dates (forecast_days business days after last date, skipping weekends and holidays)
    last_date = pd.to_datetime(time_cols[-1])
    future_dates = generate_business_dates(last_date, forecast_days, holidays)
    future_date_cols = [d.strftime('%Y-%m-%d') for d in future_dates]

    # Add new date columns to dataframe
    for date_col in future_date_cols:
        df_forecast[date_col] = 0.0

    # Create separate dataframes for upper and lower bounds (for confidence intervals)
    df_forecast_upper = df_forecast.copy()
    df_forecast_lower = df_forecast.copy()

    # Forecast each leaf node with configured model
    total_leafs = len(leaf_nodes)

    # Store detailed feature information for metadata
    feature_details_map = {}  # {leaf_id: {features, scores, cross_series_ids}}

    # Store confidence interval width for each leaf (for propagation)
    leaf_ci_widths = {}  # {leaf_id: std_error}

    for idx, leaf_id in enumerate(leaf_nodes):
        # Get model for this leaf node
        selected_model = model_config.get(leaf_id, 'Prophet')

        status_text.text(f"Forecasting {leaf_id} with {selected_model} ({idx+1}/{total_leafs})...")

        # Get historical values
        leaf_row = df[df['Row_ID'] == leaf_id]
        if len(leaf_row) > 0:
            # IMPORTANT: APUVA needs FULL historical data (2006-2025), ML models use 2019+
            if selected_model.upper() == 'APUVA':
                # APUVA: Use FULL data (2006-2025) for accurate year-over-year calculations
                values = leaf_row[time_cols].values.flatten()
                dates = pd.to_datetime(time_cols)
                external_series_data = {}  # APUVA doesn't use external features
                feature_details_map[leaf_id] = {'model': selected_model, 'uses_features': False}
            else:
                # ML Models: Use 2019+ data with external features
                time_cols_ml = [col for col in time_cols if pd.to_datetime(col) >= pd.Timestamp('2019-01-01')]
                values = leaf_row[time_cols_ml].values.flatten()
                dates = pd.to_datetime(time_cols_ml)
                external_series_data = cross_series_map.get(leaf_id, {})

                # Store feature details for ML models - WITH FEATURE SELECTION
                if selected_model.upper() in ['RANDOMFOREST', 'XGBOOST', 'LIGHTGBM']:
                    try:
                        # Run OPTIMIZED feature engineering to get actual selected features
                        from utils.feature_engineering_optimized import create_features_optimized, select_top_features_optimized
                        from scipy.stats import spearmanr

                        ts_df_temp = pd.DataFrame({'ds': dates, 'y': values})
                        train_features_temp = create_features_optimized(
                            ts_df_temp,
                            lag_steps=90,
                            holidays_list=holidays,
                            external_series=external_series_data
                        )

                        # Get ALL features and their scores
                        available_features = [col for col in train_features_temp.columns if col not in ['ds', 'date', 'value']]
                        feature_scores_all = {}

                        X_all = train_features_temp[available_features].fillna(0).replace([np.inf, -np.inf], 0)
                        y_all = train_features_temp['value'].fillna(0)

                        for col in available_features:
                            try:
                                corr, _ = spearmanr(X_all[col], y_all)
                                feature_scores_all[col] = abs(corr) if not np.isnan(corr) else 0
                            except:
                                feature_scores_all[col] = 0

                        # Select top 25 features (with volatility priority)
                        top_features, _ = select_top_features_optimized(train_features_temp, top_k=25)
                        selected_features = [f for f in top_features if f in available_features]

                        # Get scores for ALL selected features (all 25)
                        selected_feature_scores = {f: feature_scores_all.get(f, 0) for f in selected_features}

                        # Sort by score for better readability
                        selected_feature_scores = dict(sorted(selected_feature_scores.items(), key=lambda x: x[1], reverse=True))

                        feature_details_map[leaf_id] = {
                            'model': selected_model,
                            'uses_features': True,
                            'selected_features_all': list(selected_feature_scores.keys()),
                            'feature_scores_all': selected_feature_scores,
                            'total_features_selected': len(selected_features),
                            'total_features_generated': len(available_features)
                        }
                    except Exception as e:
                        # Fallback if feature selection fails
                        feature_details_map[leaf_id] = {
                            'model': selected_model,
                            'uses_features': False,
                            'error': str(e)
                        }
                else:
                    feature_details_map[leaf_id] = {'model': selected_model, 'uses_features': False}

            # Generate forecast with configured model and cross-series features
            forecast_values, forecast_dates_returned = forecast_single_leaf(
                leaf_id, values, dates, selected_model, forecast_days, future_date_cols, external_series_data
            )

            # Convert to numpy array if needed
            if hasattr(forecast_values, 'values'):
                forecast_values = forecast_values.values
            forecast_values = np.array(forecast_values)

            # Check for NaN in forecast
            if np.any(np.isnan(forecast_values)):
                st.warning(f"⚠️ {leaf_id} forecast contains NaN! Replacing with last value.")
                last_val = values[-1] if not np.isnan(values[-1]) else 0
                forecast_values = np.full(len(future_date_cols), last_val)

            # Calculate confidence intervals based on model residuals (consistent with 4_Prediksi.py)
            # This approach measures actual model error rather than historical volatility
            try:
                # Calculate in-sample predictions to get residuals
                if selected_model.upper() in ['RANDOMFOREST', 'XGBOOST', 'LIGHTGBM']:
                    # ML models: use OPTIMIZED feature engineering and get fitted values
                    from utils.feature_engineering_optimized import create_features_optimized, select_top_features_optimized

                    ts_df_temp = pd.DataFrame({'ds': dates, 'y': values})
                    train_features_temp = create_features_optimized(
                        ts_df_temp,
                        lag_steps=90,
                        holidays_list=holidays,
                        external_series=external_series_data
                    )

                    # Select features (with volatility priority)
                    available_features = [col for col in train_features_temp.columns if col not in ['ds', 'date', 'value']]
                    top_features, _ = select_top_features_optimized(train_features_temp, top_k=25)
                    feature_cols = [f for f in top_features if f in available_features]

                    if len(feature_cols) > 0:
                        X_train = train_features_temp[feature_cols].fillna(0).replace([np.inf, -np.inf], 0)
                        y_train = train_features_temp['value'].fillna(0)

                        # Train model and get fitted values
                        if selected_model.upper() == 'RANDOMFOREST':
                            from sklearn.ensemble import RandomForestRegressor
                            model_temp = RandomForestRegressor(n_estimators=100, max_depth=10, random_state=42, n_jobs=-1)
                        elif selected_model.upper() == 'XGBOOST':
                            import xgboost as xgb
                            model_temp = xgb.XGBRegressor(n_estimators=100, learning_rate=0.05, max_depth=5, random_state=42, verbosity=0)
                        elif selected_model.upper() == 'LIGHTGBM':
                            from lightgbm import LGBMRegressor
                            model_temp = LGBMRegressor(n_estimators=100, learning_rate=0.05, max_depth=5, random_state=42, verbose=-1)

                        model_temp.fit(X_train, y_train)
                        fitted_values = model_temp.predict(X_train)

                        # Calculate residuals
                        residuals = y_train.values - fitted_values
                        ci_std = np.std(residuals)
                    else:
                        # Fallback if no features
                        ci_std = np.std(values) * 0.1

                elif selected_model.upper() == 'APUVA':
                    # APUVA: use historical volatility as fallback (no easy way to get residuals)
                    ci_std = np.std(values) * 0.5  # Conservative estimate

                elif selected_model.upper() == 'PROPHET':
                    # Prophet: use historical volatility (Prophet has built-in CI, but we standardize here)
                    ci_std = np.std(values) * 0.3

                else:
                    # Default fallback
                    ci_std = np.std(values) * 0.3

            except Exception as e:
                # Fallback to conservative estimate if calculation fails
                ci_std = np.std(values) * 0.3 if len(values) > 0 else np.abs(np.mean(forecast_values)) * 0.1

            # 95% confidence interval (±1.96 * std)
            ci_multiplier = 1.96
            forecast_upper = forecast_values + (ci_multiplier * ci_std)
            forecast_lower = forecast_values - (ci_multiplier * ci_std)

            # Store CI width for this leaf (for later aggregation)
            leaf_ci_widths[leaf_id] = ci_std

            # Update dataframe with forecast values (round to 0 decimals)
            # Map forecast values to dates correctly
            leaf_idx = df_forecast[df_forecast['Row_ID'] == leaf_id].index
            if len(leaf_idx) > 0:
                # Create date -> value mapping from returned forecast
                forecast_dates_str = [d.strftime('%Y-%m-%d') if hasattr(d, 'strftime') else d for d in forecast_dates_returned]
                forecast_map = dict(zip(forecast_dates_str, forecast_values))
                forecast_upper_map = dict(zip(forecast_dates_str, forecast_upper))
                forecast_lower_map = dict(zip(forecast_dates_str, forecast_lower))

                # Assign point forecast, upper, and lower bounds
                for date_col in future_date_cols:
                    if date_col in forecast_map:
                        df_forecast.at[leaf_idx[0], date_col] = round(float(forecast_map[date_col]), 2)
                        df_forecast_upper.at[leaf_idx[0], date_col] = round(float(forecast_upper_map[date_col]), 2)
                        df_forecast_lower.at[leaf_idx[0], date_col] = round(float(forecast_lower_map[date_col]), 2)
                    else:
                        st.warning(f"⚠️ Missing forecast for {leaf_id} on {date_col}")

        # Update progress
        progress_bar.progress((idx + 1) / total_leafs)

    status_text.text("Aggregating hierarchy...")

    # Aggregate calculated nodes (point forecast, upper, and lower bounds)
    df_forecast = aggregate_hierarchy(df_forecast, future_date_cols)
    df_forecast_upper = aggregate_hierarchy(df_forecast_upper, future_date_cols)
    df_forecast_lower = aggregate_hierarchy(df_forecast_lower, future_date_cols)

    # Add D.UCL and D.LCL rows below row D
    status_text.text("Adding UCL/LCL rows for D...")

    # Find row D
    d_row = df_forecast[df_forecast['Row_ID'] == 'D']
    if len(d_row) > 0:
        # Get D upper and lower bounds
        d_upper_row = df_forecast_upper[df_forecast_upper['Row_ID'] == 'D'].copy()
        d_lower_row = df_forecast_lower[df_forecast_lower['Row_ID'] == 'D'].copy()

        # Create D.UCL row
        d_ucl_row = d_upper_row.copy()
        d_ucl_row['Row_ID'] = 'D.UCL'
        d_ucl_row['Row_Label'] = 'D. NET SUPPLY DEMAND VALAS - Upper Control Limit (95% CI)'
        d_ucl_row['Level'] = 1

        # Create D.LCL row
        d_lcl_row = d_lower_row.copy()
        d_lcl_row['Row_ID'] = 'D.LCL'
        d_lcl_row['Row_Label'] = 'D. NET SUPPLY DEMAND VALAS - Lower Control Limit (95% CI)'
        d_lcl_row['Level'] = 1

        # Insert D.UCL and D.LCL right after D
        d_idx = df_forecast[df_forecast['Row_ID'] == 'D'].index[0]

        # Insert UCL and LCL rows
        df_forecast = pd.concat([
            df_forecast.iloc[:d_idx+1],
            d_ucl_row,
            d_lcl_row,
            df_forecast.iloc[d_idx+1:]
        ], ignore_index=True)

    progress_bar.progress(1.0)
    status_text.text("✅ Forecast completed!")

    # Prepare metadata for version tracking
    timestamp = datetime.now().isoformat()

    # Get config name if using custom mode
    config_name = '-'
    if model_mode == "Custom (Hasil Evaluasi)" and len(available_configs) > 0:
        config_name = selected_config['name']

    # Calculate model usage counts
    model_counts = {}
    for model_name in model_config.values():
        model_counts[model_name] = model_counts.get(model_name, 0) + 1

    # Prepare detailed series information
    series_details = []
    for leaf_id in leaf_nodes:
        leaf_label = df[df['Row_ID'] == leaf_id]['Row_Label'].values[0] if len(df[df['Row_ID'] == leaf_id]) > 0 else leaf_id
        model_used = model_config.get(leaf_id, 'Prophet')
        feature_info = feature_details_map.get(leaf_id, {})

        series_details.append({
            'row_id': leaf_id,
            'row_label': leaf_label,
            'model': model_used,
            'uses_features': feature_info.get('uses_features', False),
            'selected_features_all': feature_info.get('selected_features_all', []),
            'feature_scores_all': feature_info.get('feature_scores_all', {}),
            'total_features_selected': feature_info.get('total_features_selected', 0),
            'total_features_generated': feature_info.get('total_features_generated', 0)
        })

    metadata = {
        'timestamp': timestamp,
        'forecast_days': forecast_days,
        'model_mode': model_mode,
        'model_config': model_config,
        'model_counts': model_counts,
        'config_name': config_name,
        'total_leaf_nodes': total_leafs,
        'future_date_cols': future_date_cols,
        'leaf_nodes': leaf_nodes,
        'time_cols': time_cols,
        'cross_series_used': needs_cross_series,
        'cross_series_top_k': 30 if needs_cross_series else 0,
        'holidays_count': len(holidays),
        'last_historical_date': time_cols[-1],
        'first_forecast_date': future_date_cols[0],
        'last_forecast_date': future_date_cols[-1],
        'series_details': series_details,  # NEW: Detailed per-series information
        'feature_details': feature_details_map  # NEW: Feature engineering details
    }

    # Save forecast version with metadata
    version_id = save_forecast_version(df_forecast, metadata)

    # Save to session state for adjustment feature
    st.session_state['df_forecast'] = df_forecast
    st.session_state['future_date_cols'] = future_date_cols
    st.session_state['leaf_nodes'] = leaf_nodes
    st.session_state['time_cols'] = time_cols
    st.session_state['forecast_version_id'] = version_id
    st.session_state['forecast_metadata'] = metadata

    st.success(f"✅ Successfully forecasted {total_leafs} leaf nodes and aggregated hierarchy!")
    st.success(f"✅ Added D.UCL and D.LCL rows (95% confidence intervals) for Net Supply Demand Valas")
    st.info(f"📦 **Versi Forecast**: `{version_id}` - Tersimpan dengan metadata lengkap")

    # Show preview
    st.subheader(f"📋 Preview Data Forecast (30 Hari Terakhir + {forecast_days} Hari Forecast)")

    # Show last 10 historical dates + forecast dates
    preview_cols = ['Row_ID', 'Row_Label', 'Level'] + time_cols[-10:] + future_date_cols
    preview_df = df_forecast[preview_cols]

    # Style forecast columns with red background
    def highlight_forecast_cols(s):
        # Create a style dataframe with empty strings
        styles = pd.DataFrame('', index=s.index, columns=s.columns)
        # Apply red background to forecast columns
        for col in future_date_cols:
            if col in styles.columns:
                styles[col] = 'background-color: #ffcccc'
        return styles

    styled_preview = preview_df.style.apply(highlight_forecast_cols, axis=None)
    st.dataframe(styled_preview, use_container_width=True, height=600)

    # ========================================================================
    # VISUALISASI NET SUPPLY DEMAND VALAS (D) WITH UCL/LCL
    # ========================================================================
    st.divider()
    st.subheader("📈 Visualisasi Net Supply Demand Valas (D) dengan Confidence Intervals")

    # Get row D, D.UCL, D.LCL
    d_row = df_forecast[df_forecast['Row_ID'] == 'D']
    d_ucl_row = df_forecast[df_forecast['Row_ID'] == 'D.UCL']
    d_lcl_row = df_forecast[df_forecast['Row_ID'] == 'D.LCL']

    if len(d_row) > 0 and len(d_ucl_row) > 0 and len(d_lcl_row) > 0:
        import plotly.graph_objects as go

        # Get last 30 historical dates + forecast dates
        historical_cols = time_cols[-30:]
        all_dates = historical_cols + future_date_cols

        # Extract values
        d_values = d_row[all_dates].values.flatten()
        d_ucl_values = d_ucl_row[all_dates].values.flatten()
        d_lcl_values = d_lcl_row[all_dates].values.flatten()

        # Convert dates to datetime (keep as strings for Plotly compatibility)
        date_objects = all_dates  # Use string dates directly instead of Timestamp objects

        # Create figure
        fig = go.Figure()

        # Split into historical and forecast for different styling
        split_idx = len(historical_cols)

        # Historical line (gray)
        fig.add_trace(go.Scatter(
            x=date_objects[:split_idx],
            y=d_values[:split_idx],
            mode='lines',
            line=dict(color='gray', width=1),
            name='Train',
            hovertemplate='Historical: %{y:,.0f}<extra></extra>'
        ))

        # Test data (actual) - using black line for last few historical points before forecast
        # This creates a visual connection between historical and forecast
        test_window = min(10, split_idx)  # Show last 10 points as "test"
        fig.add_trace(go.Scatter(
            x=date_objects[split_idx-test_window:split_idx],
            y=d_values[split_idx-test_window:split_idx],
            mode='lines',
            line=dict(color='black', width=2),
            name='Recent Historical',
            hovertemplate='Recent: %{y:,.0f}<extra></extra>'
        ))

        # Prepend last historical point for continuity (no gap)
        forecast_dates = date_objects[split_idx:]
        last_hist_date = date_objects[split_idx-1]
        last_hist_value = d_values[split_idx-1]
        last_ucl_value = d_ucl_values[split_idx-1]
        last_lcl_value = d_lcl_values[split_idx-1]

        extended_dates = [last_hist_date] + list(forecast_dates)
        extended_values = [last_hist_value] + list(d_values[split_idx:])
        extended_ucl = [last_hist_value] + list(d_ucl_values[split_idx:])
        extended_lcl = [last_hist_value] + list(d_lcl_values[split_idx:])

        # Confidence interval band (starting from last historical)
        fig.add_trace(go.Scatter(
            x=extended_dates,
            y=extended_ucl,
            mode='lines',
            line=dict(width=0),
            showlegend=False
        ))

        fig.add_trace(go.Scatter(
            x=extended_dates,
            y=extended_lcl,
            mode='lines',
            line=dict(width=0),
            fill='tonexty',
            fillcolor='rgba(68, 138, 255, 0.2)',
            name='Confidence Interval',
            hovertemplate='CI: %{y:,.0f}<extra></extra>'
        ))

        # Forecast line (red, dotted) - consistent with 4_Prediksi.py
        fig.add_trace(go.Scatter(
            x=extended_dates,
            y=extended_values,
            mode='lines',
            line=dict(color='red', dash='dot', width=2),
            name='Forecast',
            hovertemplate='Forecast: %{y:,.0f}<extra></extra>'
        ))

        # Update layout (consistent with 4_Prediksi.py)
        fig.update_layout(
            title=f"Net Supply Demand Valas (D) - {forecast_days} Days Forecast",
            xaxis_title="Date",
            yaxis_title="Value (USD Juta)",
            hovermode='x unified',
            height=600,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )

        st.plotly_chart(fig, use_container_width=True)

        # Show statistics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Last Historical Value", f"{d_values[split_idx-1]:,.0f}")
        with col2:
            forecast_mean = np.mean(d_values[split_idx:])
            st.metric("Average Forecast", f"{forecast_mean:,.0f}")
        with col3:
            ci_width = np.mean(d_ucl_values[split_idx:] - d_lcl_values[split_idx:])
            st.metric("Avg CI Width", f"±{ci_width/2:,.0f}")

    st.info("✅ Forecast berhasil di-generate! Silahkan ke halaman **⚙️ Adjustment** untuk melakukan penyesuaian manual.")

# Download section
st.divider()
st.subheader("📥 Download Forecast")

if 'df_forecast' in st.session_state:
    from io import BytesIO
    from utils.forecast_version import get_version_summary, format_timestamp

    buffer = BytesIO()
    df_forecast = st.session_state['df_forecast']
    forecast_days = len(st.session_state.get('future_date_cols', []))
    future_date_cols = st.session_state.get('future_date_cols', [])
    version_id = st.session_state.get('forecast_version_id', '')

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Sheet 1: Forecast Data
        df_forecast.to_excel(writer, index=False, sheet_name='SDV Forecast')

        # Format forecast columns with red background
        from openpyxl.styles import PatternFill
        worksheet = writer.sheets['SDV Forecast']

        # Red fill for forecast columns
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

        # Find column indices for forecast dates
        header_row = list(df_forecast.columns)
        forecast_col_indices = []
        for col_name in future_date_cols:
            if col_name in header_row:
                col_idx = header_row.index(col_name) + 1  # openpyxl uses 1-based indexing
                forecast_col_indices.append(col_idx)

        # Apply red background to forecast columns (header + all data rows)
        for col_idx in forecast_col_indices:
            # Header row
            worksheet.cell(row=1, column=col_idx).fill = red_fill
            # Data rows
            for row_idx in range(2, len(df_forecast) + 2):  # +2 because 1=header, data starts at 2
                worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

        # Sheet 2: Metadata Summary
        if version_id:
            summary = get_version_summary(version_id)
            metadata_rows = []
            metadata_rows.append(['Version ID', version_id])
            metadata_rows.append(['Timestamp', format_timestamp(summary.get('timestamp', ''))])
            metadata_rows.append(['Forecast Days', summary.get('forecast_days', 0)])
            metadata_rows.append(['Model Mode', summary.get('model_mode', '-')])
            metadata_rows.append(['Config Name', summary.get('config_name', '-')])
            metadata_rows.append(['Total Leaf Nodes', summary.get('total_leaf_nodes', 0)])
            metadata_rows.append(['Cross-Series Used', 'Yes' if summary.get('cross_series_used', False) else 'No'])
            metadata_rows.append(['Cross-Series Top K', summary.get('cross_series_top_k', 0)])
            metadata_rows.append(['Holidays Count', summary.get('holidays_count', 0)])
            metadata_rows.append(['Last Historical Date', summary.get('last_historical_date', '-')])
            metadata_rows.append(['First Forecast Date', summary.get('first_forecast_date', '-')])
            metadata_rows.append(['Last Forecast Date', summary.get('last_forecast_date', '-')])
            metadata_rows.append(['', ''])
            metadata_rows.append(['Model Usage', ''])
            for model_name, count in summary.get('model_counts', {}).items():
                metadata_rows.append([model_name, count])

            df_metadata = pd.DataFrame(metadata_rows, columns=['Parameter', 'Value'])
            df_metadata.to_excel(writer, index=False, sheet_name='Metadata')

            # Sheet 3: Series Summary
            series_details = summary.get('series_details', [])
            if len(series_details) > 0:
                series_rows = []
                for detail in series_details:
                    series_rows.append({
                        'Row_ID': detail.get('row_id', '-'),
                        'Row_Label': detail.get('row_label', '-'),
                        'Model': detail.get('model', '-'),
                        'Uses_Features': 'Yes' if detail.get('uses_features', False) else 'No',
                        'Features_Selected': detail.get('total_features_selected', 0),
                        'Features_Generated': detail.get('total_features_generated', 0)
                    })

                df_series_summary = pd.DataFrame(series_rows)
                df_series_summary.to_excel(writer, index=False, sheet_name='Series Summary')

            # Sheet 4: Feature Details (ALL 25 features per ML model series with correlation scores)
            feature_detail_rows = []
            for detail in series_details:
                if detail.get('uses_features', False):
                    row_id = detail.get('row_id', '-')
                    row_label = detail.get('row_label', '-')
                    feature_scores = detail.get('feature_scores_all', {})

                    if feature_scores:
                        for rank, (feature_name, score) in enumerate(feature_scores.items(), start=1):
                            feature_detail_rows.append({
                                'Row_ID': row_id,
                                'Row_Label': row_label,
                                'Rank': rank,
                                'Feature_Name': feature_name,
                                'Correlation_Score': round(score, 4)
                            })

            if len(feature_detail_rows) > 0:
                df_feature_details = pd.DataFrame(feature_detail_rows)
                df_feature_details.to_excel(writer, index=False, sheet_name='Feature Details')

    buffer.seek(0)

    st.download_button(
        label="📥 Download Forecast Original (dengan Metadata)",
        data=buffer,
        file_name=f"sdv_forecast_{forecast_days}days_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type="primary",
        use_container_width=True
    )

    st.caption("""
    💡 File berisi 4 sheet:
    - **SDV Forecast**: Data forecast original dengan kolom merah untuk tanggal forecast
    - **Metadata**: Informasi versi, model usage, dan parameters
    - **Series Summary**: Ringkasan per-series (model apa, berapa fitur yang dipakai)
    - **Feature Details**: SEMUA 25 fitur terpilih per series ML dengan skor korelasi Spearman (sorted by rank)
    """)
else:
    st.info("📝 Silahkan generate forecast terlebih dahulu.")

# Footer
st.divider()
st.caption("🔮 Forecast telah tersimpan di session. Lanjut ke halaman **⚙️ Adjustment** untuk penyesuaian manual.")
